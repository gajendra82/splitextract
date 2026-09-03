"""Excel (.xlsx / .xls) invoice extraction for the existing split-and-extract API.

Parses structured workbook rows into the same flat invoice dict shape that Gemini
returns, so callers can run enforce_schema() and build the Laravel JSON response
without a separate schema.
"""

from __future__ import annotations

import logging
import os
import re
from dataclasses import dataclass, field
from datetime import datetime, date, time as dt_time
from typing import Any, Dict, Iterable, Iterator, List, Optional, Tuple

logger = logging.getLogger(__name__)

PARSER_VERSION = "1.1.0"

_STOCK_HEADER_KEYS = {
    "opening",
    "stockin",
    "purchases",
    "sales",
    "closing",
    "sale",
    "packingfactor",
    "uom",
    "manufacturer",
}

# ---------------------------------------------------------------------------
# Header normalization
# ---------------------------------------------------------------------------

_CANONICAL_FIELDS: Dict[str, Tuple[str, ...]] = {
    "invoice_no": (
        "invoice no",
        "invoice number",
        "invoice num",
        "invoice #",
        "inv no",
        "inv number",
        "inv num",
        "inv #",
        "bill no",
        "bill number",
        "bill #",
        "invoiceno",
        "invoicenumber",
        "invno",
        "document no",
        "doc no",
        "voucher no",
    ),
    "invoice_date": (
        "invoice date",
        "bill date",
        "inv date",
        "date",
        "doc date",
        "document date",
        "invoicedate",
        "billing date",
    ),
    "customer": (
        "customer",
        "customer name",
        "retailer",
        "retailer name",
        "chemist",
        "chemist name",
        "buyer",
        "buyer name",
        "party",
        "party name",
        "sold to",
        "bill to",
        "to party",
        "customername",
        # GRN / hospital sales exports
        "hospital",
        "hospital name",
        "hospitalname",
        "institution",
        "institution name",
        "account name",
        "account",
        # SAP / hospital-wise sales: Card Name = hospital, Card Code = hospital code
        "card name",
        "cardname",
        "card name/bp name",
        "bp name",
    ),
    "customer_address": (
        "customer address",
        "party address",
        "buyer address",
        "retailer address",
        "address",
        "billing address",
        "ship to address",
        "bill to address",
        "delivery address",
        "shipping address",
        "hospital city",
        "hospitalcity",
        "hospital address",
    ),
    "customer_gstin": (
        "customer gstin",
        "customer gst",
        "party gstin",
        "buyer gstin",
        "retailer gstin",
        "to gstin",
    ),
    "vendor": (
        "vendor",
        "vendor name",
        "distributor",
        "distributor name",
        "stockist",
        "stockist name",
        "supplier",
        "supplier name",
        "from party",
        "seller",
        "seller name",
        "shipped unit",
        "shipping unit",
        "dispatch unit",
        "shipped from",
    ),
    "vendor_gstin": (
        "vendor gstin",
        "vendor gst",
        "distributor gstin",
        "stockist gstin",
        "supplier gstin",
        "from gstin",
        "seller gstin",
    ),
    "product_description": (
        "product",
        "product name",
        "product description",
        "item",
        "item name",
        "item description",
        "description",
        "particulars",
        "medicine",
        "medicine name",
        "drug name",
    ),
    "quantity": (
        "quantity",
        "qty",
        "qty.",
        "qnty",
        "sale qty",
        "sales qty",
        "billed qty",
        "shipment qty",
        "shipment quantity",
        "shipped qty",
        "dispatch qty",
        "dispatched qty",
        "delivery qty",
        "delivered qty",
        "received qty",
        "grn qty",
    ),
    "unit_of_measure": (
        "uom",
        "sale uom",
        "sales uom",
        "base uom",
        "unit of measure",
    ),
    "unit_price": (
        "unit price",
        "rate",
        "price",
        "ptr",
        "pts",
        "selling rate",
        "unit rate",
        "rate per unit",
    ),
    "total_amount": (
        "amount",
        "net amount",
        "line amount",
        "total amount",
        "value",
        "taxable amount",
        "taxable value",
        "line total",
        "net value",
    ),
    "tax": (
        "gst",
        "gst amount",
        "tax",
        "tax amount",
        "cgst",
        "sgst",
        "igst",
        "vat",
    ),
    "tax_amount": (
        "gst %",
        "gst%",
        "tax %",
        "tax%",
        "gst rate",
        "tax rate",
        "cgst %",
        "sgst %",
        "igst %",
    ),
    "lot_batch_number": (
        "batch",
        "batch no",
        "batch number",
        "batch #",
        "lot",
        "lot no",
        "lot number",
        "lot/batch",
        "lot batch",
        "batchno",
    ),
    "mrp": (
        "mrp",
        "m.r.p",
        "m.r.p.",
        "maximum retail price",
    ),
    "hsn_code": (
        "hsn",
        "hsn code",
        "hsn/sac",
        "sac",
        "hsn sac",
    ),
    "sku_code": (
        "sku",
        "sku code",
        "item code",
        "product code",
        "code",
        "material code",
    ),
    "free_quantity": (
        "free",
        "free qty",
        "free quantity",
        "bonus",
        "scheme qty",
    ),
    "expiry_date": (
        "expiry",
        "expiry date",
        "exp",
        "exp date",
        "exp.",
    ),
    "mfg": (
        "mfg",
        "mfg date",
        "manufacture date",
        "manufacturing date",
    ),
    "phone": (
        "phone",
        "mobile",
        "contact",
        "tel",
        "telephone",
        "phone no",
        "mobile no",
    ),
    "total": (
        "invoice total",
        "grand total",
        "net total",
        "bill total",
        "invoice amount",
        "total value",
    ),
    "irn": (
        "irn",
        "irn no",
        "irn number",
    ),
}


class HeaderNormalizer:
    """Map flexible spreadsheet headers onto canonical invoice field names."""

    def __init__(self, extra_aliases: Optional[Dict[str, Iterable[str]]] = None):
        self._alias_to_canonical: Dict[str, str] = {}
        for canonical, aliases in _CANONICAL_FIELDS.items():
            for alias in aliases:
                self._alias_to_canonical[self.normalize_key(alias)] = canonical
            self._alias_to_canonical[self.normalize_key(canonical)] = canonical
        if extra_aliases:
            for canonical, aliases in extra_aliases.items():
                for alias in aliases:
                    self._alias_to_canonical[self.normalize_key(alias)] = canonical

    @staticmethod
    def normalize_key(value: Any) -> str:
        text = str(value or "").strip().lower()
        text = text.replace("\xa0", " ")
        text = re.sub(r"[\s_\-./\\]+", "", text)
        return text

    def canonicalize(self, header: Any) -> Optional[str]:
        key = self.normalize_key(header)
        if not key:
            return None

        exact = self._alias_to_canonical.get(key)
        if exact is not None:
            return exact

        # Business rule: unit price comes ONLY from a rate/unit-price column,
        # NEVER from MRP. Tax/discount "rate" columns are also excluded.
        if "mrp" in key or "maximumretailprice" in key:
            return "mrp"
        if any(
            tok in key
            for tok in (
                "gst", "tax", "cgst", "sgst", "igst",
                "disc", "zren", "ydis", "octroi", "cess",
            )
        ):
            return None
        if "unitprice" in key or "rate" in key:
            return "unit_price"

        return None

    def map_headers(self, headers: Iterable[Any]) -> Dict[int, str]:
        """Return {column_index: canonical_field} for recognized headers."""
        mapping: Dict[int, str] = {}
        seen: Dict[str, int] = {}
        for idx, header in enumerate(headers):
            canonical = self.canonicalize(header)
            if not canonical:
                continue
            # Prefer first occurrence of a canonical field
            if canonical in seen:
                continue
            seen[canonical] = idx
            mapping[idx] = canonical
        return mapping


# ---------------------------------------------------------------------------
# Result containers
# ---------------------------------------------------------------------------


@dataclass
class ExcelParseMetadata:
    source: str = "excel"
    worksheet_name: str = ""
    original_filename: str = ""
    processing_time: float = 0.0
    total_rows: int = 0
    processed_rows: int = 0
    ignored_rows: int = 0
    failed_rows: int = 0
    parser_version: str = PARSER_VERSION
    warnings: List[str] = field(default_factory=list)

    def to_dict(self) -> Dict[str, Any]:
        return {
            "source": self.source,
            "worksheet_name": self.worksheet_name,
            "original_filename": self.original_filename,
            "processing_time": round(self.processing_time, 3),
            "total_rows": self.total_rows,
            "processed_rows": self.processed_rows,
            "ignored_rows": self.ignored_rows,
            "failed_rows": self.failed_rows,
            "parser_version": self.parser_version,
            "warnings": list(self.warnings),
        }


@dataclass
class ExcelExtractResult:
    invoices: List[Dict[str, Any]] = field(default_factory=list)
    metadata: ExcelParseMetadata = field(default_factory=ExcelParseMetadata)
    success: bool = True
    error: Optional[str] = None


# ---------------------------------------------------------------------------
# Value helpers (local copies — avoid importing app.py at module load)
# ---------------------------------------------------------------------------


def _normalize_whitespace(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\xa0", " ")
    text = re.sub(r"[\r\n\t]+", " ", text)
    text = re.sub(r"[ ]{2,}", " ", text).strip()
    return text


def _normalize_invoice_number(value: Any) -> str:
    text = _normalize_whitespace(value)
    if not text:
        return ""
    text = text.replace("£", "E").replace("€", "E").replace("$", "S")
    text = text.strip(".,;:-_ ")
    return text.upper()


def _normalize_numeric(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return ""
    if isinstance(value, (int, float)):
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return f"{value:.4f}".rstrip("0").rstrip(".") if isinstance(value, float) else str(value)
    text = _normalize_whitespace(value)
    if not text:
        return ""
    text = re.sub(r"[₹$€£,\s]", "", text)
    text = re.sub(r"[^\d.\-]", "", text)
    return text


def _normalize_phone(value: Any) -> str:
    text = _normalize_whitespace(value)
    if not text:
        return ""
    digits = re.sub(r"\D", "", text)
    return digits or text


def _normalize_date(value: Any) -> str:
    """Normalize to YYYY-MM-DD when possible; otherwise return cleaned string."""
    if value is None or value == "":
        return ""

    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, dt_time):
        return ""

    # Excel serial date (openpyxl data_only may already convert; xlrd may not)
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        try:
            # Excel epoch 1899-12-30
            from datetime import timedelta

            serial = float(value)
            if 20000 <= serial <= 80000:  # rough valid range ~1954–2119
                dt = datetime(1899, 12, 30) + timedelta(days=serial)
                return dt.strftime("%Y-%m-%d")
        except Exception:
            pass
        return _normalize_numeric(value)

    text = _normalize_whitespace(value)
    if not text:
        return ""

    formats = (
        "%Y-%m-%d",
        "%d-%m-%Y",
        "%d/%m/%Y",
        "%d.%m.%Y",
        "%d %b %Y",
        "%d-%b-%Y",
        "%d %B %Y",
        "%Y/%m/%d",
        "%m/%d/%Y",
        "%d-%m-%y",
        "%d/%m/%y",
    )
    candidates = [text]
    if "\\" in text:
        candidates.append(text.replace("\\", "/"))
    for candidate in candidates:
        for fmt in formats:
            try:
                return datetime.strptime(candidate, fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue
    return text


def _cell_to_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return _normalize_whitespace(value)


def _is_empty_row(values: Iterable[Any]) -> bool:
    for v in values:
        if v is None:
            continue
        if isinstance(v, str) and not v.strip():
            continue
        return False
    return True


def _composite_group_key(mapped: Dict[str, Any]) -> str:
    """Identify one document when the sheet carries no invoice number column.

    GRN / dispatch exports repeat the consignment header (vendor, customer,
    address, date) on every product row, so rows sharing that header describe
    one delivery rather than one invoice each. Returns "" when the row has too
    little header data to group on, in which case it stays standalone.
    """
    parts = [
        _cell_to_str(mapped.get(field_name)).upper()
        for field_name in ("vendor", "customer", "customer_address", "invoice_date")
    ]
    customer, invoice_date = parts[1], parts[3]
    if not customer or not invoice_date:
        return ""
    return "|".join(parts)


def _row_has_product_signal(mapped: Dict[str, Any]) -> bool:
    product = _normalize_whitespace(mapped.get("product_description"))
    qty = _normalize_whitespace(mapped.get("quantity"))
    amount = _normalize_whitespace(mapped.get("total_amount"))
    return bool(product or qty or amount)


# ---------------------------------------------------------------------------
# Sheet / workbook readers
# ---------------------------------------------------------------------------


def _score_header_row(cells: List[Any], normalizer: HeaderNormalizer) -> int:
    mapping = normalizer.map_headers(cells)
    score = len(mapping)
    # Prefer rows that look like invoice tables
    if "invoice_no" in mapping.values():
        score += 3
    if "product_description" in mapping.values():
        score += 2
    if "quantity" in mapping.values():
        score += 1
    return score


def _pick_header_row(
    sample_rows: List[Tuple[int, List[Any]]],
    normalizer: HeaderNormalizer,
) -> Tuple[Optional[int], Dict[int, str], List[Any]]:
    best_idx: Optional[int] = None
    best_map: Dict[int, str] = {}
    best_headers: List[Any] = []
    best_score = 0
    for row_idx, cells in sample_rows:
        score = _score_header_row(cells, normalizer)
        if score > best_score:
            best_score = score
            best_idx = row_idx
            best_map = normalizer.map_headers(cells)
            best_headers = cells
    if best_score < 2:
        return None, {}, []
    return best_idx, best_map, best_headers


def _is_stock_statement_sheet(headers: List[Any], col_map: Dict[int, str]) -> bool:
    """Skip opening/purchases/closing stock grids — they are not POD/GRN invoices."""
    mapped = set(col_map.values())
    if "invoice_no" in mapped or "customer" in mapped:
        return False
    stock_hits = 0
    for header in headers:
        key = HeaderNormalizer.normalize_key(header)
        if not key:
            continue
        if (
            key in _STOCK_HEADER_KEYS
            or key.startswith("opening")
            or key.startswith("closing")
            or key.startswith("stock")
        ):
            stock_hits += 1
    return stock_hits >= 3


def _hospital_pivot_columns(headers: List[Any], col_map: Dict[int, str]) -> List[int]:
    """Unmapped columns whose headers look like hospital / account names."""
    mapped_idxs = set(col_map.keys())
    hospital_cols: List[int] = []
    for idx, header in enumerate(headers):
        if idx in mapped_idxs:
            continue
        text = _normalize_whitespace(header)
        if len(text) < 4:
            continue
        key = HeaderNormalizer.normalize_key(text)
        if (
            key in _STOCK_HEADER_KEYS
            or key.startswith("opening")
            or key.startswith("closing")
            or key.startswith("stock")
        ):
            continue
        hospital_cols.append(idx)
    return hospital_cols


def _is_hospital_qty_pivot(headers: List[Any], col_map: Dict[int, str]) -> bool:
    """Item rows × hospital-name columns with quantities (no invoice/customer col)."""
    mapped = set(col_map.values())
    if "invoice_no" in mapped or "customer" in mapped:
        return False
    if _is_stock_statement_sheet(headers, col_map):
        return False
    if "product_description" not in mapped and "sku_code" not in mapped:
        return False
    return len(_hospital_pivot_columns(headers, col_map)) >= 3


def invoices_are_usable(invoices: List[Dict[str, Any]]) -> bool:
    """True when at least one invoice has a hospital/customer or a real invoice no."""
    for invoice in invoices:
        customer = _normalize_whitespace(invoice.get("customer"))
        invoice_no = str(invoice.get("invoice_no") or "").strip()
        items = invoice.get("line_items") or []
        if not items:
            continue
        if customer:
            return True
        if invoice_no and not invoice_no.upper().startswith("UNKNOWN"):
            return True
    return False


def _iter_xlsx_rows(path: str) -> Tuple[str, Iterator[Tuple[int, List[Any], bool]]]:
    """Yield (1-based_row_index, values, is_hidden) from first usable worksheet."""
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True, keep_vba=False)
    try:
        if not wb.sheetnames:
            raise ValueError("No worksheets found in workbook")

        # Prefer first visible sheet with content
        chosen = None
        for name in wb.sheetnames:
            ws = wb[name]
            # Skip obviously empty sheets by peeking first rows later
            chosen = ws
            break
        if chosen is None:
            raise ValueError("No worksheets found in workbook")

        sheet_name = chosen.title
        logger.info("Excel workbook loaded; using worksheet '%s'", sheet_name)

        def _gen() -> Iterator[Tuple[int, List[Any], bool]]:
            try:
                for row_idx, row in enumerate(
                    chosen.iter_rows(values_only=False), start=1
                ):
                    values = []
                    row_num = row_idx
                    for c in row:
                        values.append(getattr(c, "value", None))
                        # Prefer real cell.row when available (EmptyCell has none)
                        cell_row = getattr(c, "row", None)
                        if isinstance(cell_row, int) and cell_row > 0:
                            row_num = cell_row
                    # Hidden row detection (best-effort; limited in read_only)
                    is_hidden = False
                    try:
                        dim = chosen.row_dimensions.get(row_num)
                        if dim is not None and getattr(dim, "hidden", False):
                            is_hidden = True
                    except Exception:
                        is_hidden = False
                    yield row_num, values, is_hidden
            finally:
                wb.close()

        return sheet_name, _gen()
    except Exception:
        try:
            wb.close()
        except Exception:
            pass
        raise


def _iter_xls_rows(path: str) -> Tuple[str, Iterator[Tuple[int, List[Any], bool]]]:
    import xlrd

    wb = xlrd.open_workbook(path, on_demand=True)
    if wb.nsheets < 1:
        raise ValueError("No worksheets found in workbook")

    sheet = wb.sheet_by_index(0)
    sheet_name = sheet.name
    logger.info("Excel (.xls) workbook loaded; using worksheet '%s'", sheet_name)

    def _gen() -> Iterator[Tuple[int, List[Any], bool]]:
        try:
            for r in range(sheet.nrows):
                values = []
                for c in range(sheet.ncols):
                    cell = sheet.cell(r, c)
                    val = cell.value
                    # Convert xlrd dates
                    if cell.ctype == xlrd.XL_CELL_DATE:
                        try:
                            tup = xlrd.xldate_as_tuple(val, wb.datemode)
                            if tup[0] > 0:
                                val = datetime(*tup[:6])
                            else:
                                val = dt_time(tup[3], tup[4], tup[5])
                        except Exception:
                            pass
                    values.append(val)
                yield r + 1, values, False
        finally:
            try:
                wb.release_resources()
            except Exception:
                pass

    return sheet_name, _gen()


def _open_row_iterator(path: str, ext: str) -> Tuple[str, Iterator[Tuple[int, List[Any], bool]]]:
    ext = ext.lower()
    if ext == ".xlsx":
        return _iter_xlsx_rows(path)
    if ext == ".xls":
        # Prefer xlrd for legacy .xls; fall back to openpyxl if somehow xlsx mislabeled
        try:
            return _iter_xls_rows(path)
        except Exception as xls_err:
            logger.warning("xlrd failed for %s (%s); trying openpyxl", path, xls_err)
            return _iter_xlsx_rows(path)
    raise ValueError(f"Unsupported Excel extension: {ext}")


def _xlsx_sheet_rows(ws) -> List[Tuple[int, List[Any], bool]]:
    rows: List[Tuple[int, List[Any], bool]] = []
    for row_idx, row in enumerate(ws.iter_rows(values_only=False), start=1):
        values = []
        row_num = row_idx
        for cell in row:
            values.append(getattr(cell, "value", None))
            cell_row = getattr(cell, "row", None)
            if isinstance(cell_row, int) and cell_row > 0:
                row_num = cell_row
        is_hidden = False
        try:
            dim = ws.row_dimensions.get(row_num)
            if dim is not None and getattr(dim, "hidden", False):
                is_hidden = True
        except Exception:
            is_hidden = False
        rows.append((row_num, values, is_hidden))
    return rows


def _xls_sheet_rows(sheet, datemode: int) -> List[Tuple[int, List[Any], bool]]:
    import xlrd

    rows: List[Tuple[int, List[Any], bool]] = []
    for r in range(sheet.nrows):
        values = []
        for c in range(sheet.ncols):
            cell = sheet.cell(r, c)
            val = cell.value
            if cell.ctype == xlrd.XL_CELL_DATE:
                try:
                    tup = xlrd.xldate_as_tuple(val, datemode)
                    if tup[0] > 0:
                        val = datetime(*tup[:6])
                    else:
                        val = dt_time(tup[3], tup[4], tup[5])
                except Exception:
                    pass
            values.append(val)
        rows.append((r + 1, values, False))
    return rows


def _load_all_sheets(path: str, ext: str) -> List[Tuple[str, List[Tuple[int, List[Any], bool]]]]:
    """Load every worksheet so GRN workbooks with MAY/JUNE (or SALE+STATEMENT) are all seen."""
    ext = ext.lower()
    sheets: List[Tuple[str, List[Tuple[int, List[Any], bool]]]] = []
    if ext == ".xlsx":
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True, keep_vba=False)
        try:
            if not wb.sheetnames:
                raise ValueError("No worksheets found in workbook")
            for name in wb.sheetnames:
                sheets.append((name, _xlsx_sheet_rows(wb[name])))
        finally:
            wb.close()
        return sheets
    if ext == ".xls":
        import xlrd

        try:
            wb = xlrd.open_workbook(path, on_demand=True)
        except Exception as xls_err:
            logger.warning("xlrd failed for %s (%s); trying openpyxl", path, xls_err)
            return _load_all_sheets(path, ".xlsx")
        try:
            if wb.nsheets < 1:
                raise ValueError("No worksheets found in workbook")
            for idx in range(wb.nsheets):
                sheet = wb.sheet_by_index(idx)
                sheets.append((sheet.name, _xls_sheet_rows(sheet, wb.datemode)))
        finally:
            try:
                wb.release_resources()
            except Exception:
                pass
        return sheets
    raise ValueError(f"Unsupported Excel extension: {ext}")


def _qty_from_cell(value: Any) -> Optional[float]:
    text = _normalize_numeric(value)
    if not text:
        return None
    try:
        qty = float(text)
    except ValueError:
        return None
    if qty == 0:
        return None
    return qty


def _unpivot_hospital_qty_sheet(
    rows: List[Tuple[int, List[Any], bool]],
    headers: List[Any],
    col_map: Dict[int, str],
    header_row_idx: int,
    worksheet_name: str,
    metadata: ExcelParseMetadata,
    unknown_start: int,
) -> List[Dict[str, Any]]:
    """Turn Item × Hospital quantity columns into one invoice per hospital."""
    hospital_cols = _hospital_pivot_columns(headers, col_map)
    groups: Dict[str, Dict[str, Any]] = {}
    order: List[str] = []
    unknown_counter = unknown_start

    for row_num, values, is_hidden in rows:
        metadata.total_rows += 1
        if row_num <= header_row_idx or is_hidden or _is_empty_row(values):
            metadata.ignored_rows += 1
            continue
        mapped = _map_row(values, col_map)
        product = _normalize_whitespace(mapped.get("product_description"))
        sku = _normalize_whitespace(mapped.get("sku_code"))
        if not product and not sku:
            metadata.ignored_rows += 1
            continue
        emitted = False
        for col_idx in hospital_cols:
            if col_idx >= len(values):
                continue
            qty = _qty_from_cell(values[col_idx])
            if qty is None:
                continue
            hospital = _normalize_whitespace(headers[col_idx] if col_idx < len(headers) else "")
            if not hospital:
                continue
            key = hospital.upper()
            if key not in groups:
                unknown_counter += 1
                invoice = _empty_invoice_shell(f"UNKNOWN_{unknown_counter}")
                invoice["customer"] = hospital
                groups[key] = invoice
                order.append(key)
            item = _build_line_item(
                {
                    "product_description": product,
                    "sku_code": sku,
                    "quantity": qty,
                    "unit_price": mapped.get("unit_price"),
                    "total_amount": mapped.get("total_amount"),
                    "lot_batch_number": mapped.get("lot_batch_number"),
                    "hsn_code": mapped.get("hsn_code"),
                    "unit_of_measure": mapped.get("unit_of_measure"),
                    "mrp": mapped.get("mrp"),
                    "expiry_date": mapped.get("expiry_date"),
                },
                row_num,
            )
            if item:
                groups[key]["line_items"].append(item)
                emitted = True
        if emitted:
            metadata.processed_rows += 1
        else:
            metadata.ignored_rows += 1

    return [_finalize_invoice(groups[key], worksheet_name) for key in order]


def _extract_invoices_from_sheet(
    rows: List[Tuple[int, List[Any], bool]],
    worksheet_name: str,
    metadata: ExcelParseMetadata,
    unknown_start: int,
) -> List[Dict[str, Any]]:
    normalizer = HeaderNormalizer()
    sample: List[Tuple[int, List[Any]]] = []
    for row_num, values, is_hidden in rows:
        if is_hidden or _is_empty_row(values):
            continue
        sample.append((row_num, values))
        if len(sample) >= 40:
            break

    header_row_idx, col_map, headers = _pick_header_row(sample, normalizer)
    if header_row_idx is None or not col_map:
        metadata.warnings.append(f"Skipped sheet '{worksheet_name}': no usable headers")
        return []

    if _is_stock_statement_sheet(headers, col_map):
        metadata.warnings.append(
            f"Skipped stock-statement sheet '{worksheet_name}'"
        )
        logger.info("Skipping stock-statement worksheet '%s'", worksheet_name)
        return []

    if _is_hospital_qty_pivot(headers, col_map):
        logger.info(
            "Hospital quantity pivot detected on '%s' (%s hospital columns)",
            worksheet_name,
            len(_hospital_pivot_columns(headers, col_map)),
        )
        return _unpivot_hospital_qty_sheet(
            rows,
            headers,
            col_map,
            header_row_idx,
            worksheet_name,
            metadata,
            unknown_start,
        )

    logger.info(
        "Headers detected on '%s' row %s: %s",
        worksheet_name,
        header_row_idx,
        sorted(set(col_map.values())),
    )
    invoices = group_rows_into_invoices(
        iter(rows),
        col_map,
        header_row_idx,
        worksheet_name,
        metadata,
        headers=headers,
    )
    shifted = unknown_start
    for invoice in invoices:
        invoice_no = str(invoice.get("invoice_no") or "")
        if invoice_no.upper().startswith("UNKNOWN"):
            shifted += 1
            invoice["invoice_no"] = f"UNKNOWN_{shifted}"
    return invoices


def _merge_sheet_invoices(
    all_invoices: List[Dict[str, Any]],
    sheet_invoices: List[Dict[str, Any]],
) -> None:
    index = {
        str(inv.get("invoice_no") or "").upper(): inv
        for inv in all_invoices
        if inv.get("invoice_no")
        and not str(inv.get("invoice_no")).upper().startswith("UNKNOWN")
    }
    for invoice in sheet_invoices:
        key = str(invoice.get("invoice_no") or "").upper()
        if key and not key.startswith("UNKNOWN") and key in index:
            index[key]["line_items"].extend(invoice.get("line_items") or [])
            _finalize_invoice(
                index[key],
                (index[key].get("_excel") or {}).get("worksheet_name") or "",
            )
            continue
        all_invoices.append(invoice)
        if key and not key.startswith("UNKNOWN"):
            index[key] = invoice


def workbook_to_tsv_for_llm(
    path: str,
    original_filename: Optional[str] = None,
    max_rows_per_sheet: int = 250,
    max_chars: int = 40000,
) -> str:
    """Compact TSV of non-stock sheets for Gemini fallback (PDF-text analogue)."""
    filename = original_filename or os.path.basename(path)
    ext = os.path.splitext(filename.lower())[1] or os.path.splitext(path.lower())[1]
    chunks: List[str] = [f"SOURCE FILE: {filename}"]
    try:
        sheets = _load_all_sheets(path, ext)
    except Exception as exc:
        return f"SOURCE FILE: {filename}\nERROR: {exc}"

    normalizer = HeaderNormalizer()
    for sheet_name, rows in sheets:
        sample = [
            (row_num, values)
            for row_num, values, hidden in rows[:40]
            if not hidden and not _is_empty_row(values)
        ]
        _, col_map, headers = _pick_header_row(sample, normalizer)
        if headers and _is_stock_statement_sheet(headers, col_map or {}):
            continue
        chunks.append(f"\n--- SHEET: {sheet_name} ---")
        emitted = 0
        for row_num, values, is_hidden in rows:
            if is_hidden or _is_empty_row(values):
                continue
            cells = [_cell_to_str(v) for v in values]
            chunks.append("\t".join(cells))
            emitted += 1
            if emitted >= max_rows_per_sheet:
                chunks.append("[... rows truncated ...]")
                break
    text = "\n".join(chunks)
    if len(text) > max_chars:
        text = text[:max_chars] + "\n[... truncated ...]"
    return text


# ---------------------------------------------------------------------------
# Grouping + invoice building
# ---------------------------------------------------------------------------


def _excel_col_letter(idx: int) -> str:
    """0-based column index -> spreadsheet letter (0->A, 26->AA)."""
    try:
        from openpyxl.utils import get_column_letter

        return get_column_letter(idx + 1)
    except Exception:
        result = ""
        n = idx
        while n >= 0:
            result = chr(ord("A") + (n % 26)) + result
            n = n // 26 - 1
        return result


def _log_unit_price_source(
    mapped: Dict[str, Any],
    values: List[Any],
    col_map: Dict[int, str],
    worksheet_name: Optional[str],
    row_num: Optional[int],
    headers: Optional[List[Any]],
) -> None:
    """Diagnostic: log which Excel cell became unit_price for the target line item."""
    if "unit_price" not in mapped:
        return
    product = str(mapped.get("product_description") or "").upper()
    batch = str(mapped.get("lot_batch_number") or "").upper()
    invoice = str(mapped.get("invoice_no") or "").upper()
    if not ("AZTREO" in product or "AZA25002" in batch or "INV26-37264" in invoice):
        return
    up_idx = next((i for i, f in col_map.items() if f == "unit_price"), None)
    if up_idx is None:
        return
    header_txt = ""
    if headers is not None and up_idx < len(headers):
        header_txt = str(headers[up_idx] or "")
    raw_value = values[up_idx] if up_idx < len(values) else None
    logger.info(
        "[unit_price source] sheet=%r excel_row=%s col=%s header=%r field=unit_price "
        "raw_cell_value=%r | invoice=%r product=%r batch=%r",
        worksheet_name,
        row_num,
        _excel_col_letter(up_idx),
        header_txt,
        raw_value,
        mapped.get("invoice_no"),
        mapped.get("product_description"),
        mapped.get("lot_batch_number"),
    )


def _map_row(
    values: List[Any],
    col_map: Dict[int, str],
    *,
    worksheet_name: Optional[str] = None,
    row_num: Optional[int] = None,
    headers: Optional[List[Any]] = None,
) -> Dict[str, Any]:
    mapped: Dict[str, Any] = {}
    for idx, field in col_map.items():
        if idx >= len(values):
            continue
        mapped[field] = values[idx]
    _log_unit_price_source(mapped, values, col_map, worksheet_name, row_num, headers)
    return mapped


def _build_line_item(mapped: Dict[str, Any], row_number: int) -> Optional[Dict[str, Any]]:
    product = _normalize_whitespace(mapped.get("product_description"))
    qty_raw = mapped.get("quantity")
    amount_raw = mapped.get("total_amount")

    if not product and not _normalize_whitespace(qty_raw) and not _normalize_whitespace(amount_raw):
        return None

    qty = _normalize_numeric(qty_raw)
    unit_price = _normalize_numeric(mapped.get("unit_price"))
    total_amount = _normalize_numeric(amount_raw)
    free_qty = _normalize_numeric(mapped.get("free_quantity"))

    # Handle qty like "22+2"
    if isinstance(qty_raw, str) and "+" in qty_raw and not free_qty:
        parts = qty_raw.split("+", 1)
        left = re.search(r"\d+(?:\.\d+)?", parts[0] or "")
        right = re.search(r"\d+(?:\.\d+)?", parts[1] or "")
        if left:
            qty = left.group(0)
        if right:
            free_qty = right.group(0)

    additional: Dict[str, Any] = {}
    mrp = _normalize_numeric(mapped.get("mrp"))
    if mrp:
        additional["mrp"] = mrp
    if free_qty:
        additional["free_quantity"] = free_qty
    expiry = _normalize_date(mapped.get("expiry_date"))
    if expiry:
        additional["expiry_date"] = expiry
    mfg = _normalize_date(mapped.get("mfg")) if mapped.get("mfg") else _normalize_whitespace(mapped.get("mfg"))
    if mfg:
        additional["mfg"] = mfg

    item: Dict[str, Any] = {
        "product_description": product,
        "quantity": qty,
        "unit_price": unit_price,
        "total_amount": total_amount,
        "hsn_code": _normalize_whitespace(mapped.get("hsn_code")),
        "lot_batch_number": _normalize_whitespace(mapped.get("lot_batch_number")),
        "sku_code": _normalize_whitespace(mapped.get("sku_code")),
        "tax_amount": _normalize_numeric(mapped.get("tax_amount")),
        "unit_of_measure": _normalize_whitespace(mapped.get("unit_of_measure")),
        "discount": "",
        "additional_fields": additional,
        "confidence": 1.0,
        "row_number": row_number,
    }
    return item


def _empty_invoice_shell(invoice_no: str) -> Dict[str, Any]:
    return {
        "invoice_no": invoice_no,
        "invoice_date": "",
        "invoice_date_raw": "",
        "vendor": "",
        "vendor_gstin": "",
        "customer": "",
        "customer_address": "",
        "customer_gstin": "",
        "tax": "",
        "total": "",
        "irn": "",
        "line_items": [],
        "confidence": 1.0,
        "_excel": {
            "source_rows": [],
            "worksheet_name": "",
        },
    }


def _apply_header_fields(invoice: Dict[str, Any], mapped: Dict[str, Any]) -> None:
    inv_no = _normalize_invoice_number(mapped.get("invoice_no"))
    if inv_no and not str(invoice.get("invoice_no", "")).startswith("UNKNOWN"):
        # Keep existing group key; header may reinforce values
        pass
    elif inv_no:
        invoice["invoice_no"] = inv_no

    date_raw = mapped.get("invoice_date")
    if date_raw not in (None, ""):
        raw_str = _cell_to_str(date_raw)
        iso = _normalize_date(date_raw)
        if iso and not invoice.get("invoice_date"):
            invoice["invoice_date"] = iso
            invoice["invoice_date_raw"] = raw_str

    for field_name in (
        "vendor",
        "vendor_gstin",
        "customer",
        "customer_address",
        "customer_gstin",
        "irn",
    ):
        val = mapped.get(field_name)
        if val in (None, ""):
            continue
        text = _normalize_whitespace(val)
        if field_name.endswith("gstin"):
            text = re.sub(r"[^A-Za-z0-9]", "", text).upper()
        if text and not invoice.get(field_name):
            invoice[field_name] = text

    tax = _normalize_numeric(mapped.get("tax"))
    if tax and not invoice.get("tax"):
        invoice["tax"] = tax

    total = _normalize_numeric(mapped.get("total"))
    if not total:
        total = _normalize_numeric(mapped.get("total_amount")) if not mapped.get("product_description") else ""
    if total and not invoice.get("total"):
        invoice["total"] = total

    phone = _normalize_phone(mapped.get("phone"))
    if phone:
        invoice.setdefault("_excel", {})["phone"] = phone


def _finalize_invoice(invoice: Dict[str, Any], worksheet_name: str) -> Dict[str, Any]:
    items = invoice.get("line_items") or []
    # If total missing, sum line amounts
    if not invoice.get("total"):
        total = 0.0
        any_amount = False
        for item in items:
            try:
                amt = float(_normalize_numeric(item.get("total_amount")) or 0)
                if amt:
                    total += amt
                    any_amount = True
            except Exception:
                continue
        if any_amount:
            invoice["total"] = f"{total:.2f}"

    excel_meta = invoice.get("_excel") or {}
    excel_meta["worksheet_name"] = worksheet_name
    invoice["_excel"] = excel_meta

    # Ensure confidence on invoice
    invoice["confidence"] = 1.0
    return invoice


def group_rows_into_invoices(
    rows: Iterator[Tuple[int, List[Any], bool]],
    col_map: Dict[int, str],
    header_row_idx: int,
    worksheet_name: str,
    metadata: ExcelParseMetadata,
    headers: Optional[List[Any]] = None,
) -> List[Dict[str, Any]]:
    """Group consecutive / matching invoice_no rows into invoice dicts."""
    invoices_by_key: Dict[str, Dict[str, Any]] = {}
    order: List[str] = []
    unknown_counter = 0
    unknown_key_by_header: Dict[str, str] = {}
    current_key: Optional[str] = None

    for row_num, values, is_hidden in rows:
        metadata.total_rows += 1

        if row_num <= header_row_idx:
            metadata.ignored_rows += 1
            continue
        if is_hidden:
            metadata.ignored_rows += 1
            metadata.warnings.append(f"Skipped hidden row {row_num}")
            continue
        if _is_empty_row(values):
            metadata.ignored_rows += 1
            continue

        try:
            mapped = _map_row(
                values,
                col_map,
                worksheet_name=worksheet_name,
                row_num=row_num,
                headers=headers,
            )
            inv_no = _normalize_invoice_number(mapped.get("invoice_no"))

            if inv_no:
                key = inv_no
            elif _row_has_product_signal(mapped) and current_key and not current_key.startswith("UNKNOWN_"):
                # Continuation row without invoice number — attach to current invoice
                key = current_key
            elif _row_has_product_signal(mapped) or any(
                _normalize_whitespace(mapped.get(f))
                for f in ("customer", "vendor", "invoice_date", "total")
            ):
                header_key = _composite_group_key(mapped)
                existing_key = unknown_key_by_header.get(header_key) if header_key else None
                if existing_key:
                    key = existing_key
                else:
                    unknown_counter += 1
                    key = f"UNKNOWN_{unknown_counter}"
                    if header_key:
                        unknown_key_by_header[header_key] = key
            else:
                metadata.ignored_rows += 1
                continue

            if key not in invoices_by_key:
                invoices_by_key[key] = _empty_invoice_shell(key)
                order.append(key)

            invoice = invoices_by_key[key]
            _apply_header_fields(invoice, mapped)

            # Ensure invoice_no on known keys
            if not key.startswith("UNKNOWN"):
                invoice["invoice_no"] = key

            item = _build_line_item(mapped, row_num)
            if item:
                invoice["line_items"].append(item)

            excel_meta = invoice.setdefault("_excel", {"source_rows": [], "worksheet_name": worksheet_name})
            excel_meta.setdefault("source_rows", []).append(row_num)

            current_key = key
            metadata.processed_rows += 1
        except Exception as row_err:
            metadata.failed_rows += 1
            metadata.warnings.append(f"Failed row {row_num}: {row_err}")
            logger.warning("Excel row %s failed: %s", row_num, row_err)
            continue

    result = [_finalize_invoice(invoices_by_key[k], worksheet_name) for k in order]
    logger.info(
        "Excel grouping complete: invoices=%s products=%s processed=%s ignored=%s failed=%s",
        len(result),
        sum(len(inv.get("line_items") or []) for inv in result),
        metadata.processed_rows,
        metadata.ignored_rows,
        metadata.failed_rows,
    )
    return result


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def extract_invoices_from_excel(
    path: str,
    original_filename: Optional[str] = None,
) -> ExcelExtractResult:
    """
    Parse an Excel workbook into flat invoice dicts compatible with enforce_schema().

    Walks every worksheet. Stock-statement grids are skipped. Hospital-name
    quantity pivots are unpivoted. Does not call Gemini/OCR — structured cells
    get confidence=1.0.
    """
    started = datetime.now()
    filename = original_filename or os.path.basename(path)
    ext = os.path.splitext(filename.lower())[1] or os.path.splitext(path.lower())[1]
    metadata = ExcelParseMetadata(original_filename=filename, worksheet_name="")
    result = ExcelExtractResult(metadata=metadata)

    if ext not in {".xlsx", ".xls"}:
        result.success = False
        result.error = f"Unsupported Excel format: {ext or '(none)'}"
        metadata.warnings.append(result.error)
        return result

    try:
        file_size = os.path.getsize(path)
        if file_size > 200 * 1024 * 1024:
            metadata.warnings.append(
                f"Large workbook ({file_size / (1024 * 1024):.1f} MB); streaming parse enabled"
            )
        logger.info(
            "Loading Excel workbook '%s' (%.2f MB, ext=%s)",
            filename,
            file_size / (1024 * 1024),
            ext,
        )

        sheets = _load_all_sheets(path, ext)
        all_invoices: List[Dict[str, Any]] = []
        unknown_start = 0
        used_sheets: List[str] = []

        for sheet_name, rows in sheets:
            if not rows:
                metadata.warnings.append(f"Skipped empty sheet '{sheet_name}'")
                continue
            sheet_invoices = _extract_invoices_from_sheet(
                rows,
                sheet_name,
                metadata,
                unknown_start,
            )
            for invoice in sheet_invoices:
                invoice_no = str(invoice.get("invoice_no") or "")
                if invoice_no.upper().startswith("UNKNOWN"):
                    try:
                        unknown_start = max(
                            unknown_start,
                            int(invoice_no.split("_", 1)[1]),
                        )
                    except (IndexError, ValueError):
                        unknown_start += 1
            if sheet_invoices:
                used_sheets.append(sheet_name)
                _merge_sheet_invoices(all_invoices, sheet_invoices)

        metadata.worksheet_name = ",".join(used_sheets) if used_sheets else (
            sheets[0][0] if sheets else ""
        )
        result.invoices = all_invoices
        if not all_invoices:
            result.success = False
            result.error = result.error or "No usable headers or invoice rows detected in workbook"
            metadata.warnings.append(result.error)
            logger.warning("Excel parse produced zero invoices for '%s'", filename)

    except Exception as exc:
        msg = str(exc).lower()
        if "password" in msg or "encrypted" in msg or "workbook is encrypted" in msg:
            result.error = "Password protected workbook"
        elif (
            "not a zip" in msg
            or "badzipfile" in msg
            or "corrupt" in msg
            or "damaged" in msg
            or "ole2" in msg
        ):
            result.error = "Corrupted or unsupported workbook"
        else:
            result.error = f"Failed to parse Excel workbook: {exc}"
        result.success = False
        metadata.warnings.append(result.error)
        logger.error("Excel extract failed for '%s': %s", filename, result.error)

    metadata.processing_time = (datetime.now() - started).total_seconds()

    try:
        import psutil

        rss_mb = psutil.Process(os.getpid()).memory_info().rss / (1024 * 1024)
        logger.info(
            "Excel parse finished in %.2fs; rss=%.1fMB invoices=%s warnings=%s",
            metadata.processing_time,
            rss_mb,
            len(result.invoices),
            len(metadata.warnings),
        )
    except Exception:
        logger.info(
            "Excel parse finished in %.2fs; invoices=%s warnings=%s",
            metadata.processing_time,
            len(result.invoices),
            len(metadata.warnings),
        )

    return result


def build_excel_ocr_text(invoice: Dict[str, Any]) -> str:
    """Build a printable text block (stored as ocr_text for schema compatibility)."""
    lines = [
        f"SOURCE: EXCEL",
        f"INVOICE NO: {invoice.get('invoice_no', '')}",
        f"INVOICE DATE: {invoice.get('invoice_date', '')}",
        f"VENDOR: {invoice.get('vendor', '')}",
        f"VENDOR GSTIN: {invoice.get('vendor_gstin', '')}",
        f"CUSTOMER: {invoice.get('customer', '')}",
        f"CUSTOMER ADDRESS: {invoice.get('customer_address', '')}",
        f"CUSTOMER GSTIN: {invoice.get('customer_gstin', '')}",
        f"TAX: {invoice.get('tax', '')}",
        f"TOTAL: {invoice.get('total', '')}",
        f"IRN: {invoice.get('irn', '')}",
        "LINE ITEMS:",
    ]
    for idx, item in enumerate(invoice.get("line_items") or [], start=1):
        af = item.get("additional_fields") or {}
        lines.append(
            f"{idx}. {item.get('product_description', '')} | "
            f"Qty={item.get('quantity', '')} | "
            f"Rate={item.get('unit_price', '')} | "
            f"Amt={item.get('total_amount', '')} | "
            f"Batch={item.get('lot_batch_number', '')} | "
            f"HSN={item.get('hsn_code', '')} | "
            f"MRP={af.get('mrp', '')}"
        )
    return "\n".join(lines)
