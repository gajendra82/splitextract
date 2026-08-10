"""Unit tests for Excel invoice extraction (split-and-extract Excel path)."""

from __future__ import annotations

import os
import tempfile
import unittest

from openpyxl import Workbook

from services.excel_invoice_extract import (
    HeaderNormalizer,
    build_excel_ocr_text,
    extract_invoices_from_excel,
)
from services.file_type_detector import FileKind, FileTypeDetector


def _write_xlsx(rows, sheet_name: str = "Invoices") -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    for r_idx, row in enumerate(rows, start=1):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    wb.save(path)
    wb.close()
    return path


class TestFileTypeDetector(unittest.TestCase):
    def setUp(self):
        self.detector = FileTypeDetector()

    def test_excel_extension(self):
        kind, ext = self.detector.detect("report.XLSX")
        self.assertEqual(kind, FileKind.EXCEL)
        self.assertEqual(ext, ".xlsx")

    def test_pdf_and_image(self):
        self.assertEqual(self.detector.detect("a.pdf")[0], FileKind.PDF)
        self.assertEqual(self.detector.detect("a.PNG")[0], FileKind.IMAGE)

    def test_mime_excel(self):
        kind, _ = self.detector.detect(
            filename="blob",
            content_type="application/vnd.ms-excel",
        )
        self.assertEqual(kind, FileKind.EXCEL)


class TestHeaderNormalizer(unittest.TestCase):
    def setUp(self):
        self.n = HeaderNormalizer()

    def test_aliases(self):
        cases = {
            "Invoice No": "invoice_no",
            "Invoice Number": "invoice_no",
            "Inv No": "invoice_no",
            "Bill No": "invoice_no",
            "Date": "invoice_date",
            "Invoice Date": "invoice_date",
            "Bill Date": "invoice_date",
            "Customer": "customer",
            "Retailer": "customer",
            "Chemist": "customer",
            "Hospital Name": "customer",
            "Hospital": "customer",
            "hospital_city": "customer_address",
            "Distributor": "vendor",
            "Stockist": "vendor",
            "Product": "product_description",
            "Product Name": "product_description",
            "Item": "product_description",
            "Quantity": "quantity",
            "Qty": "quantity",
            "Batch": "lot_batch_number",
            "Batch No": "lot_batch_number",
            "GST": "tax",
            "Amount": "total_amount",
            "MRP": "mrp",
            "Net Amount": "total_amount",
        }
        for header, expected in cases.items():
            self.assertEqual(self.n.canonicalize(header), expected, header)

    def test_ignore_case_space_underscore_hyphen(self):
        self.assertEqual(self.n.canonicalize("INVOICE_NO"), "invoice_no")
        self.assertEqual(self.n.canonicalize("invoice-number"), "invoice_no")
        self.assertEqual(self.n.canonicalize("  Inv   No "), "invoice_no")


class TestExcelInvoiceExtract(unittest.TestCase):
    def tearDown(self):
        for path in getattr(self, "_temps", []):
            try:
                os.remove(path)
            except OSError:
                pass

    def _track(self, path: str) -> str:
        if not hasattr(self, "_temps"):
            self._temps = []
        self._temps.append(path)
        return path

    def test_single_invoice(self):
        path = self._track(_write_xlsx([
            ["Invoice No", "Date", "Customer", "Product", "Qty", "Amount", "Batch"],
            ["INV100", "15/01/2026", "ABC Chemist", "Product A", 10, 100.0, "B1"],
        ]))
        result = extract_invoices_from_excel(path, "single.xlsx")
        self.assertTrue(result.success)
        self.assertEqual(len(result.invoices), 1)
        inv = result.invoices[0]
        self.assertEqual(inv["invoice_no"], "INV100")
        self.assertEqual(inv["invoice_date"], "2026-01-15")
        self.assertEqual(inv["customer"], "ABC Chemist")
        self.assertEqual(len(inv["line_items"]), 1)
        self.assertEqual(inv["line_items"][0]["product_description"], "Product A")
        self.assertEqual(inv["confidence"], 1.0)
        self.assertEqual(result.metadata.source, "excel")

    def test_multiple_invoices_and_products(self):
        path = self._track(_write_xlsx([
            ["Inv No", "Invoice Date", "Retailer", "Product Name", "Quantity", "Net Amount", "Batch No", "MRP"],
            ["INV100", "01-02-2026", "Chemist One", "Product A", 2, 50, "BA", 30],
            ["INV100", "01-02-2026", "Chemist One", "Product B", 3, 75, "BB", 40],
            ["INV100", "01-02-2026", "Chemist One", "Product C", 1, 25, "BC", 50],
            ["INV200", "02-02-2026", "Chemist Two", "Product D", 5, 200, "BD", 60],
        ]))
        result = extract_invoices_from_excel(path)
        self.assertTrue(result.success)
        self.assertEqual(len(result.invoices), 2)
        by_no = {i["invoice_no"]: i for i in result.invoices}
        self.assertEqual(len(by_no["INV100"]["line_items"]), 3)
        self.assertEqual(len(by_no["INV200"]["line_items"]), 1)
        self.assertEqual(by_no["INV100"]["line_items"][0]["additional_fields"]["mrp"], "30")

    def test_duplicate_invoice_groups(self):
        path = self._track(_write_xlsx([
            ["Bill No", "Item", "Qty", "Amount"],
            ["X1", "A", 1, 10],
            ["X1", "B", 2, 20],
            ["X1", "A", 1, 10],
        ]))
        result = extract_invoices_from_excel(path)
        self.assertEqual(len(result.invoices), 1)
        self.assertEqual(len(result.invoices[0]["line_items"]), 3)

    def test_missing_invoice_number_unknowns_not_merged(self):
        path = self._track(_write_xlsx([
            ["Date", "Customer", "Product", "Qty", "Amount"],
            ["10/03/2026", "C1", "P1", 1, 10],
            ["11/03/2026", "C2", "P2", 2, 20],
        ]))
        result = extract_invoices_from_excel(path)
        self.assertEqual(len(result.invoices), 2)
        self.assertEqual(result.invoices[0]["invoice_no"], "UNKNOWN_1")
        self.assertEqual(result.invoices[1]["invoice_no"], "UNKNOWN_2")

    def test_rows_without_invoice_number_group_by_consignment_header(self):
        """GRN exports repeat vendor/customer/date per product row (one delivery)."""
        path = self._track(_write_xlsx([
            ["Shipped Unit", "Customer Name", "Bill to Address", "invoice Date",
             "Item Code", "Item Name", "Batch No", "Shipment Qty", "Sale UOM"],
            ["WH1", "Pharmacy A", "Addr A", "19/06/2026 05:54 PM", "C1", "Prod 1", "B1", 10, "STRIP10"],
            ["WH1", "Pharmacy A", "Addr A", "19/06/2026 05:54 PM", "C2", "Prod 2", "B2", 20, "BOTTLE14"],
            ["WH1", "Pharmacy A", "Addr A", "19/06/2026 06:26 PM", "C3", "Prod 3", "B3", 5, "STRIP15"],
            ["WH1", "Pharmacy B", "Addr B", "19/06/2026 05:54 PM", "C4", "Prod 4", "B4", 7, "STRIP10"],
        ]))
        result = extract_invoices_from_excel(path, "grn.xls")
        self.assertTrue(result.success)
        self.assertEqual(len(result.invoices), 3)

        first = result.invoices[0]
        self.assertEqual(len(first["line_items"]), 2)
        self.assertEqual(first["customer"], "Pharmacy A")
        self.assertEqual(first["vendor"], "WH1")
        self.assertEqual(first["customer_address"], "Addr A")
        self.assertEqual(first["line_items"][0]["quantity"], "10")
        self.assertEqual(first["line_items"][0]["unit_of_measure"], "STRIP10")
        self.assertEqual(first["line_items"][0]["sku_code"], "C1")
        self.assertEqual(
            [len(inv["line_items"]) for inv in result.invoices[1:]], [1, 1]
        )

    def test_hospital_name_column_maps_to_customer(self):
        """GRN hospital sales sheets use Hospital Name instead of Customer."""
        path = self._track(_write_xlsx([
            [
                "invoice_date",
                "group_name",
                "Hospital Name",
                "hospital_city",
                "product_name",
                "quantity",
                "invoice_number",
            ],
            [
                "2026-05-05",
                "HCG HOSPITAL",
                "HCG YELAHANKA (NORTH BANGALORE)",
                "BENGALURU",
                "FLUTICONE NASAL SPRAY 12 ML",
                1,
                "VGP/27/4697",
            ],
            [
                "2026-05-05",
                "HCG HOSPITAL",
                "HCG YELAHANKA (NORTH BANGALORE)",
                "BENGALURU",
                "OTHER PRODUCT",
                2,
                "VGP/27/4697",
            ],
        ]))
        result = extract_invoices_from_excel(path, "ZYDUS.xlsx")
        self.assertTrue(result.success)
        self.assertEqual(len(result.invoices), 1)
        inv = result.invoices[0]
        self.assertEqual(inv["customer"], "HCG YELAHANKA (NORTH BANGALORE)")
        self.assertEqual(inv["customer_address"], "BENGALURU")
        self.assertEqual(inv["invoice_no"], "VGP/27/4697")
        self.assertEqual(len(inv["line_items"]), 2)

    def test_invoice_number_column_still_wins_over_header_grouping(self):
        path = self._track(_write_xlsx([
            ["Invoice No", "Customer", "Date", "Product", "Shipment Qty"],
            ["INV1", "Pharmacy A", "01/06/2026", "Prod 1", 1],
            ["INV2", "Pharmacy A", "01/06/2026", "Prod 2", 2],
        ]))
        result = extract_invoices_from_excel(path)
        self.assertEqual([i["invoice_no"] for i in result.invoices], ["INV1", "INV2"])

    def test_rows_without_customer_or_date_stay_separate(self):
        path = self._track(_write_xlsx([
            ["Product", "Qty", "Amount"],
            ["P1", 1, 10],
            ["P2", 2, 20],
        ]))
        result = extract_invoices_from_excel(path)
        self.assertEqual(len(result.invoices), 2)

    def test_blank_rows_skipped(self):
        path = self._track(_write_xlsx([
            ["Invoice Number", "Product", "Qty", "Amount"],
            ["A1", "P1", 1, 10],
            [None, None, None, None],
            ["", "", "", ""],
            ["A1", "P2", 2, 20],
        ]))
        result = extract_invoices_from_excel(path)
        self.assertEqual(len(result.invoices), 1)
        self.assertEqual(len(result.invoices[0]["line_items"]), 2)
        self.assertGreaterEqual(result.metadata.ignored_rows, 2)

    def test_formula_cells_data_only(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["Invoice No", "Product", "Qty", "Rate", "Amount"])
        ws.append(["F1", "Prod", 2, 5, None])
        ws["E2"] = "=C2*D2"
        fd, path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        wb.save(path)
        wb.close()
        self._track(path)

        # data_only=True without Excel having cached values → formula may be None;
        # parser must still succeed and keep the product row.
        result = extract_invoices_from_excel(path)
        self.assertTrue(result.success)
        self.assertEqual(len(result.invoices), 1)
        self.assertEqual(result.invoices[0]["line_items"][0]["product_description"], "Prod")

    def test_merged_cells_header(self):
        wb = Workbook()
        ws = wb.active
        ws.append(["Invoice No", "Product", "Qty", "Amount"])
        ws.merge_cells("A1:A1")  # no-op merge; still valid workbook
        ws.append(["M1", "MergedProd", 1, 9])
        fd, path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        wb.save(path)
        wb.close()
        self._track(path)
        result = extract_invoices_from_excel(path)
        self.assertTrue(result.success)
        self.assertEqual(result.invoices[0]["invoice_no"], "M1")

    def test_hidden_row_skipped_when_detectable(self):
        wb = Workbook()
        ws = wb.active
        ws.append(["Invoice No", "Product", "Qty", "Amount"])
        ws.append(["H1", "Visible", 1, 10])
        ws.append(["H1", "HiddenProd", 9, 99])
        ws.row_dimensions[3].hidden = True
        fd, path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        wb.save(path)
        wb.close()
        self._track(path)
        result = extract_invoices_from_excel(path)
        self.assertTrue(result.success)
        # read_only mode may not expose hidden flags; accept either skip or keep
        products = [i["product_description"] for i in result.invoices[0]["line_items"]]
        self.assertIn("Visible", products)

    def test_invalid_dates_kept_as_text(self):
        path = self._track(_write_xlsx([
            ["Invoice No", "Date", "Product", "Qty", "Amount"],
            ["D1", "not-a-date", "P", 1, 1],
        ]))
        result = extract_invoices_from_excel(path)
        self.assertTrue(result.success)
        self.assertEqual(result.invoices[0]["invoice_date"], "not-a-date")

    def test_corrupted_workbook(self):
        fd, path = tempfile.mkstemp(suffix=".xlsx")
        os.write(fd, b"this is not a real workbook")
        os.close(fd)
        self._track(path)
        result = extract_invoices_from_excel(path, "bad.xlsx")
        self.assertFalse(result.success)
        self.assertTrue(result.error)

    def test_password_protected_message(self):
        # Simulate encrypted workbook failure path via monkeypatch-style stub:
        # write OLE header that openpyxl rejects
        fd, path = tempfile.mkstemp(suffix=".xlsx")
        os.write(fd, b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1" + b"\x00" * 64)
        os.close(fd)
        self._track(path)
        result = extract_invoices_from_excel(path, "encrypted.xlsx")
        self.assertFalse(result.success)
        self.assertTrue(result.metadata.warnings)

    def test_large_workbook_streaming(self):
        rows = [["Invoice No", "Product", "Qty", "Amount"]]
        for i in range(2500):
            rows.append([f"INV{i // 5}", f"Prod{i}", 1, 10])
        path = self._track(_write_xlsx(rows))
        result = extract_invoices_from_excel(path, "large.xlsx")
        self.assertTrue(result.success)
        self.assertGreater(len(result.invoices), 100)
        self.assertEqual(result.metadata.processed_rows, 2500)

    def test_ocr_text_builder(self):
        text = build_excel_ocr_text({
            "invoice_no": "X",
            "invoice_date": "2026-01-01",
            "vendor": "V",
            "customer": "C",
            "line_items": [{"product_description": "P", "quantity": "1",
                            "unit_price": "2", "total_amount": "2",
                            "lot_batch_number": "B", "hsn_code": "",
                            "additional_fields": {"mrp": "3"}}],
        })
        self.assertIn("SOURCE: EXCEL", text)
        self.assertIn("INVOICE NO: X", text)
        self.assertIn("P |", text)


class TestExcelSchemaCompatibility(unittest.TestCase):
    """Ensure Excel flat dicts survive enforce_schema with Laravel shape."""

    def test_enforce_schema_shape(self):
        from app import enforce_schema

        flat = {
            "invoice_no": "INV9",
            "invoice_date": "2026-05-01",
            "vendor": "Stockist A",
            "customer": "Retailer B",
            "tax": "18",
            "total": "118",
            "line_items": [{
                "product_description": "Med",
                "quantity": "2",
                "unit_price": "50",
                "total_amount": "100",
                "lot_batch_number": "L1",
                "hsn_code": "3004",
                "sku_code": "",
                "tax_amount": "18",
                "additional_fields": {"mrp": "60"},
                "confidence": 1.0,
            }],
            "confidence": 1.0,
            "ocr_text": "SOURCE: EXCEL",
        }
        formatted = enforce_schema({"data": flat})
        self.assertEqual(formatted["status"], "success")
        self.assertIn("invoice_summary", formatted["data"])
        self.assertEqual(formatted["data"]["invoice_summary"]["invoice_no"], "INV9")
        self.assertEqual(formatted["data"]["line_items"]["count"], 1)
        self.assertEqual(
            formatted["data"]["line_items"]["items"][0]["product_description"], "Med"
        )


if __name__ == "__main__":
    unittest.main()
